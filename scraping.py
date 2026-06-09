import streamlit as st
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import random
import os
import time
import re
import subprocess
from urllib.parse import urlparse
from datetime import datetime

# ----------------------------
# PAGE TITLE
# ----------------------------
st.set_page_config(page_title="Lead Scrapper Tool")

# st.title("Lead Scrapper Tool")

# ----------------------------
# SESSION STATE
# ----------------------------
if "stop_scraping" not in st.session_state:
    st.session_state.stop_scraping = False

    # ----------------------------
    # URL INPUT
    # ----------------------------
url = st.text_input(
    "Enter URL",
    placeholder="https://www.justdial.com/Chennai/Clinics/nct-10101647"
    )

    # ----------------------------
    # BUTTONS
    # ----------------------------
col1, col2 = st.columns(2)

with col1:
    start_button = st.button("Start Scraping")

with col2:
    stop_button = st.button("Stop Scraping")
# ----------------------------
# STOP BUTTON
# ----------------------------
 # 
if stop_button:
    st.session_state.stop_scraping = True
    st.warning("Scraping stopped by user")
    st.stop()
if start_button:
    st.session_state.stop_scraping = False

if not start_button:
    st.stop()          

if not url:
    st.error("Please enter URL")
    st.stop()               

           

output = subprocess.check_output(
 r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version',
     shell=True
    ).decode()

version = re.search(r'(\d+)\.', output).group(1)

driver = uc.Chrome(
version_main=int(version),
use_subprocess=True
 )
processed_urls = set()

try:

    driver.get(url)

    st.write("Page Title :", driver.title)
    # st.write("Current URL :", driver.current_url)

    time.sleep(5)


    # SCROLL PAGE
     # ----------------------------
    last_height = driver.execute_script(
    "return document.body.scrollHeight"
        )
    before = driver.execute_script("return window.pageYOffset")

    driver.execute_script("window.scrollBy(0,1000)")
    time.sleep(2)

    # after = driver.execute_script("return window.pageYOffset")

    # st.write("Before:", before)
    # st.write("After :", after)
    # st.write(last_height)
    for i in range(5):
        if st.session_state.stop_scraping:
        
            driver.quit()
            st.warning("Scraping stopped")
            st.stop()

        st.write(f"Scrolling : {i + 1}")

        driver.execute_script(
        "window.scrollTo(0, document.body.scrollHeight);"
        )

        time.sleep(random.randint(5, 10))

    last_count = 0

    scroll_no = 0

    while True:
        scroll_no += 1

    # Scroll slowly
        for _ in range(10):
            driver.execute_script(
                "window.scrollBy(0, 500);"
            )
            time.sleep(1)
        height = driver.execute_script(
        "return document.body.scrollHeight"
        )

        # Wait for new records
        time.sleep(5)

        cards = driver.find_elements(
            By.CSS_SELECTOR,
            "div.resultbox"
        )
        # for idx, card in enumerate(cards[:3]):
        #     st.write(
        #         idx,
        #         card.get_attribute("id")
        #     )
        # st.write(
        # f"Scroll {i+1} | Height={height} | Cards={len(cards)}"
        #     )
        current_count = len(cards)
        # st.write("current_count",current_count)

        st.write(f"Records Loaded: {current_count}")

        if current_count == last_count:
            st.write("Waiting for more records...")
            time.sleep(10)

            cards = driver.find_elements(
                By.CSS_SELECTOR,
                "div.resultbox"
            )   
            

            if len(cards) == last_count:
                st.write("crads",len(cards))
                st.write("All records loaded")
                break

        last_count = current_count

    
        

    soup = BeautifulSoup(driver.page_source, "html.parser")
    business_links = set()
    all_links = soup.find_all("a", href=True)

    for link in all_links:
        href = link["href"]
      
        if "/Chennai/" in href and "nct-" not in href:
            full_url = href
            if full_url.startswith("/"):
                full_url = "https://www.justdial.com" + full_url
                # st.write("frist if",full_url)
            if full_url not in business_links:
                # st.write("second if")
                business_links.add(full_url)
                st.write("Total business links :", len(business_links))
    if len(business_links) == 0:
        st.error("No business links found")
        driver.quit()
        st.stop()


    path_parts = url.split("/")
    category_name = path_parts[-2]
    excel_file = f"{category_name}.xlsx"
    st.write("Category:", category_name)
    st.write("Excel File:", excel_file)
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
        st.write("Existing Excel file opened")

    else:
        st.write("new  Excel file opened")
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Business Name",
            "Website",
            "Justdial URL",
            "Phone Number",
            "Address",
            "Date",
            ])
    wb.save(excel_file)

    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_urls.add(str(row[2]).strip())
    #st.write("New Excel file created")
    st.write("Excel Path :", os.path.abspath(excel_file)) 
    progress_bar = st.progress(0)
    status_text = st.empty()

    # OPEN BUSINESS PAGES
    count = 0
    for business_url in business_links:
        if st.session_state.stop_scraping:
            driver.quit()
            st.warning("Scraping stopped")
            st.stop()
        count += 1
        progress = count / len(business_links)
        progress_bar.progress(progress)
        status_text.write(
        f"Processing {count} / {len(business_links)}"
        )
   
    # st.warning("Scraping stopped")
            # SKIP DUPLICATE----------------------------
        if business_url in processed_urls:
            continue
        
            # ----------------------------
        if count % 20 == 0:
            st.write("Restarting browser...")
            driver.quit()
            time.sleep(5)
            driver = uc.Chrome(version_main=int(version),
            use_subprocess=True
            )
            time.sleep(5)
        try:
            driver.get(business_url)
            WebDriverWait(driver, 15).until(
            EC.presence_of_element_located(
            (By.TAG_NAME, "title")
            ))
            detail_soup = BeautifulSoup(
            driver.page_source,
            "html.parser"
            )

            # ----------------------------
            # BUSINESS NAME
            # ----------------------------
            business_name = (
            driver.title.split("in")[0].strip()
            )
            detail_links = detail_soup.find_all(
            "a",
            href=True
            )

            phone_number = "Not Found"

            for dlink in detail_links:
                text = dlink.text.strip()
                if text.isdigit() and len(text) >= 10:
                    phone_number = text
                    break
    
        
            address = "No Address"
            address_tag = detail_soup.find(
            "a",
            class_="color111"
            )
            if address_tag:
                address = address_tag.get_text(
                " ",
                strip=True
                )
            website = "No Website"

            for dlink in detail_links:
                dhref = dlink["href"].lower()
                if (
                    "http" in dhref
                    and "justdial" not in dhref
                    and "facebook" not in dhref
                    and "instagram" not in dhref
                    and "youtube" not in dhref
                    and "whatsapp" not in dhref
                    ):
                    website = dhref
                    break

                                     
            # driver.get(business_url)
            # WebDriverWait(driver, 15).until(
            # EC.presence_of_element_located(
            # (By.TAG_NAME, "title")
            # )
            # )
            clean_url = driver.current_url.split("?")[0]
            detail_soup = BeautifulSoup(
            driver.page_source,
            "html.parser"
            )

            if clean_url in existing_urls:
                st.write(f"Skipping duplicate: {clean_url}")
                continue

            Date = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
            )                                                                                
            ws.append([
                business_name,
                website,
                clean_url,
                phone_number,
                address,
                Date,
                ])
            existing_urls.add(clean_url)
            wb.save(excel_file)
            processed_urls.add(clean_url)
        except Exception as e:
            st.error(f"Error : {e}")
    driver.quit()

    st.success("Scraping Completed")
    st.write("Excel saved :", excel_file)
    if os.path.exists(excel_file):
         with open(excel_file, "rb") as file:
            st.download_button(
            label="Download Excel",
            data=file,
            file_name=excel_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
             )
    st.write(f"Loaded {len(existing_urls)} existing URLs")
except Exception as e:
    st.error(f"Main Error : {e}")

try:
    driver.quit()
except:
    pass

                             